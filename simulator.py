"""
Simulator class

Donghun Lee 2018
"""


from copy import deepcopy
from statistics import mean

import numpy as np
import time
from openpyxl import Workbook

from auction import Auction
import sim_lib as sl


class Simulator:
    """
    Simulates ad-click auction over time
    """

    def __init__(self, randseed=12345):
        """ initializes the auction simulating environment

        :param randseed: seed for the simulator. it is used for click and conversion sampling and tiebreaking
        """
        self.time_last = time.time()
        self.prng = np.random.RandomState(randseed)
        self.pols, self.puids = None, None
        self.possible_bids = list(range(10))
        self.auctions = None
        self.attrs = []
        self.t = 0
        self.max_t = None
        self.hist = []
        self.events = []
        self.p_infos = {}
        self.time_spent = {}
        self._time_log('simulator')

    def _time_log(self, n):
        t = time.time()
        if n in self.time_spent.keys():
            self.time_spent[n] += t - self.time_last
        else:
            self.time_spent[n] = t - self.time_last
        self.time_last = t

    def read_in_auction(self, aucts):
        """ reads in output from Auction.generate_sample(), and initializes policies

        :param aucts: output from Auction class. The maximum simulate-able iterations depends on this
        :return: none. After running this, it is possible to run the simulation
        """
        self.time_last = time.time()

        self.auctions = aucts
        self.attrs = sorted(list(set([a['attr'] for a in aucts])))
        self.max_t = self.auctions[-1]['iter']
        self._init_pols()

        self._time_log('simulator')


    def _init_pols(self):
        """
        internal function. Loads and initializes policies
        :return:
        """
        self.time_last = time.time()

        if self.pols is None and self.puids is None and self.attrs != []:
            self.pols, self.puids = sl.load_policies(self.attrs, self.possible_bids, self.max_t)
            self.p_infos = {ix: [] for ix in range(len(self.pols))}
            for puid in self.puids:
                self.time_spent[puid] = 0.0
        else:
            pass

        self._time_log('simulator')

    def get_num_clicks(self, bid, auct):
        theta = {'a': auct['theta'],
                 'bid': 0,                 # TODO: adjust this with attribute as well?
                 '0': 0,                   # TODO: baseline click probability
                 'max_click_prob': 0.5}
        p_click = sl.get_click_prob(theta, bid)
        num_clicks = self.prng.binomial(auct['num_auct'], p_click)
        return num_clicks, p_click

    def step(self):
        """
        simulates one timestep in the auction
        :return: True if auction is simulated, False if no auction data is present
        """
        self.time_last = time.time()

        self.t += 1
        auction_happened = False
        events = []

        for a in self.auctions:
            if a['iter'] != self.t:
                continue

            if len(self.hist) == 0:
                costs_sum = [0.0] * len(self.pols)
                revenues_sum = [0.0] * len(self.pols)
                profits_sum = [0.0] * len(self.pols)
            else:
                costs_sum = deepcopy(self.hist[-1]['costs_cumulative'])
                revenues_sum = deepcopy(self.hist[-1]['revenues_cumulative'])
                profits_sum = deepcopy(self.hist[-1]['profits_cumulative'])

            auction_happened = True
            # bids = [float(p.bid(a['attr'])) for p in self.pols]
            bids = []
            for puid, p in zip(self.puids, self.pols):
                self._time_log('simulator')
                this_bid = p.bid(a['attr'])
                self._time_log(puid)
                bids.append(float(this_bid))

            max_bid_pols_ix = sl.max_ix(bids)
            winning_bid = bids[max_bid_pols_ix[0]]
            num_clicks, p_click = self.get_num_clicks(winning_bid, a)
            cost = sl.compute_second_price_cost(bids, size=num_clicks)  # second price auction
            conversion = Auction.get_conversion(a['prob_conversion'], self.prng, size=num_clicks)
            revenue = Auction.get_revenue_sample(a['avg_revenue'], self.prng, size=num_clicks)

            winning_pol_ix = []
            for ix in range(num_clicks):
                winner_ix = int(self.prng.choice(max_bid_pols_ix))
                winning_pol_ix.append(winner_ix)
                costs_sum[winner_ix] += 1 * cost[ix]
                revenues_sum[winner_ix] += conversion[ix] * revenue[ix]
                profits_sum[winner_ix] = revenues_sum[winner_ix] - costs_sum[winner_ix]
                event = {'iter': self.t,
                         'attr': a['attr'],
                         'auctions_in_iter': a['num_auct'],
                         'bids': bids,
                         'winning_pol_id': winner_ix,
                         'winning_pol_name': self.puids[winner_ix],
                         'winning_bid': winning_bid,
                         'num_click': 1,
                         'cost_per_click': cost[ix],
                         'num_conversion': conversion[ix],
                         'revenue_per_conversion': conversion[ix] * revenue[ix],
                         'costs_cumulative': deepcopy(costs_sum),
                         'revenues_cumulative': deepcopy(revenues_sum),
                         'profits_cumulative': deepcopy(profits_sum)}
                events.append(event)
                self.events.append(event)
            if num_clicks == 0:
                winner_ix = int(self.prng.choice(max_bid_pols_ix))
                event = {'iter': self.t,
                         'attr': a['attr'],
                         'auctions_in_iter': a['num_auct'],
                         'bids': bids,
                         'winning_pol_id': winner_ix,
                         'winning_pol_name': self.puids[winner_ix],
                         'winning_bid': winning_bid,
                         'num_click': 0,
                         'cost_per_click': '',
                         'num_conversion': 0,
                         'revenue_per_conversion': '',
                         'costs_cumulative': deepcopy(costs_sum),
                         'revenues_cumulative': deepcopy(revenues_sum),
                         'profits_cumulative': deepcopy(profits_sum)}
                events.append(event)
                self.events.append(event)

            # keep aggregate history for output
            auct_res = deepcopy(a)
            auct_res['bids'] = bids
            auct_res['num_click'] = num_clicks
            auct_res['p_click'] = p_click
            auct_res['cost_per_click'] = mean(cost) if num_clicks > 0 else ''
            auct_res['num_conversion'] = sum(conversion)
            auct_res['revenue_per_conversion'] = sum([ncr * rpc for (ncr, rpc) in zip(conversion, revenue)])/sum(conversion) if sum(conversion) > 0 else ''
            auct_res['costs_cumulative'] = deepcopy(costs_sum)
            auct_res['revenues_cumulative'] = deepcopy(revenues_sum)
            auct_res['profits_cumulative'] = deepcopy(profits_sum)
            auct_res['time_spent'] = deepcopy(self.time_spent)
            self.hist.append(auct_res)
            # end of auction events handling

        # if nothing happened. this is the way to go
        if len(events) == 0:
            return auction_happened

        # aggregate information over one iteration is assembled, for each policy
        p_infos = {}

        for p_ix, p in enumerate(self.pols):
            p_infos[p_ix] = []
            last_a = events[0]['attr']
            bunch = []

            if len(self.p_infos[p_ix]) == 0:
                profit_sum = 0.0
            else:
                profit_sum = self.p_infos[p_ix][-1][-1]['your_profit_cumulative']

            for ev in events + [{'attr': 'guard_dummy'}]:
                this_a = ev['attr']
                if last_a == this_a and ev != events[-1]:
                    bunch.append(ev)
                else:
                    # aggregate information
                    p_info = {'iter': bunch[0]['iter'],
                              'attr': bunch[0]['attr'],
                              'num_auct': bunch[0]['auctions_in_iter'],
                              'your_bid': bunch[0]['bids'][p_ix],
                              'winning_bid': bunch[0]['winning_bid']}
                    impressions = []
                    clicks = []
                    costs = []
                    conversions = []
                    revenues = []
                    for ev_b in bunch:
                        if p_ix == ev_b['winning_pol_id']:
                            impressions.append(1)   # TODO: bug on this? #impr == #clicks  always??
                            clicks.append(ev_b['num_click'])
                            costs.append(ev_b['cost_per_click'])
                            conversions.append(ev_b['num_conversion'])
                            revenues.append(ev_b['revenue_per_conversion'])

                    p_add_info = {'num_impression': sum(impressions),
                                  'num_click': sum(clicks),
                                  'cost_per_click': mean(costs) if sum(clicks) > 0 else '',
                                  'num_conversion': sum(conversions),
                                  'revenue_per_conversion': mean(revenues) if sum(conversions) > 0 else ''}

                    this_cost = sum([c if c != '' else 0 for c in costs])
                    this_revenue = sum([ncr * rpc for (ncr, rpc) in zip(conversions, [r if r != '' else 0 for r in revenues])])
                    profit_sum += this_revenue - this_cost
                    p_info['your_profit_cumulative'] = profit_sum
                    p_info.update(p_add_info)
                    p_infos[p_ix].append(p_info)

                    # refresh rolling aggregator and its guard
                    bunch = [ev]
                    last_a = this_a

        # post-auction learning session for policies
        for p_ix, p in enumerate(self.pols):
            self._time_log('simulator')
            p.learn(p_infos[p_ix])
            self._time_log(self.puids[p_ix])

            self.p_infos[p_ix].append(p_infos[p_ix])

        # finish up
        self._time_log('simulator')
        print(self.time_spent)

        return auction_happened

    def output_hist_to_xlsx(self, fname, pol_name=None):
        """
        outputs aggregate history to xslx
        :param fname: output file name
        :param pol_name: if not provided, prints master version.
        :return:
        """
        wb = Workbook()
        ws = wb.active
        outs = ['iter', 'attr', 'lambda', 'num_auct', 'bids', 'theta', 'p_click', 'num_click', 'cost_per_click',
                'num_conversion', 'revenue_per_conversion', 'costs_cumulative', 'revenues_cumulative', 'profits_cumulative']
        ws.append(outs),
        for h in self.hist:
            ws.append([str(h[k]) if isinstance(h[k], (list, tuple)) else h[k] for k in outs])
            # ws.append([str(['{:.2f}'.format(i) for i in h[k] ]) if isinstance(h[k], (list, tuple)) else '{:.2f}'.format(h[k]) for k in outs])
        wb.save(fname)

    def output_time_logged_to_xlsx(self, fname):
        wb = Workbook()
        ws = wb.active
        outs = ['simulator'] + self.puids
        ws.append(['t_{}'.format(k) for k in outs])
        for h in self.hist:
            ws.append([h['time_spent'][k] for k in outs])
        wb.save(fname)

    def output_policy_info_to_xlsx(self, fname, pol_ix):
        """
        outputs information given to policy for each iteration

        :param fname:
        :param pol_ix:
        :return:
        """

        wb = Workbook()
        ws = wb.active
        outs = ['iter', 'attr', 'num_auct', 'your_bid', 'winning_bid', 'num_impression', 'num_click', 'cost_per_click',
                'num_conversion', 'revenue_per_conversion', 'your_profit_cumulative']
        ws.append(outs)
        for p_info_iter in self.p_infos[pol_ix]:
            for p_info in p_info_iter:
                ws.append([p_info[k] if not isinstance(p_info[k], tuple) else str(p_info[k]) for k in outs])
        wb.save(fname)


    def output_all(self):
        self.output_hist_to_xlsx("output_master_aggregate.xlsx")
        for ix, puid in enumerate(self.puids):
            self.output_policy_info_to_xlsx("output_policy_info_{}.xlsx".format(puid), ix)
        self.output_time_logged_to_xlsx("output_time_spent_in_seconds.xlsx")


if __name__ == "__main__":
    t_start = time.time()
    print("{:.2f} sec: start loading simulator".format(time.time() - t_start))
    sim = Simulator()

    param, attrs = Auction.read_init_xlsx("auction_ini_02.xlsx")
    auc = Auction(param, attrs)
    aucts = auc.generate_sample()
    # aucts = sl.load_auction_p("auction_01.p")   # loading from a snapshot example.
    sim.read_in_auction(aucts)

    print("{:.2f} sec: finished loading simulator".format(time.time() - t_start))
    for t in range(param['max iteration']):
        sim_res = sim.step()
        print("{:.2f} sec: simulation iter {}, auction happened? {}".format(time.time() - t_start, t, sim_res))

    sim.output_all()
    print("{:.2f} sec: created output files.".format(time.time() - t_start))
    pass