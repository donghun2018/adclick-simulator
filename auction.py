"""
Auction class

Donghun Lee 2018
"""

from itertools import product

import numpy as np

from openpyxl import load_workbook


class Auction:
    """ Class to generate auction arrivals

    """
    @staticmethod
    def read_init_xlsx(fname):
        wb = load_workbook("auction_ini_01.xlsx")
        ws = wb.active

        param = {}
        attrs = []
        for r in ws.iter_rows():
            cell1 = r[0].value
            if cell1 == "END_OF_CONFIGURATION":
                break
            elif cell1 is None:
                attrs.append({})
                continue
            if cell1.split('_')[0] == "attribute":
                attr = attrs[-1]
                attr[cell1.split('_')[1]] = r[1].value
            else:
                param[cell1] = r[1].value
        return param,attrs

    def __init__(self, init_param, attributes):
        self.prng = np.random.RandomState(init_param['random seed'])
        self.max_t = int(init_param['max iteration'])
        self.attrs = self._parse_attributes(attributes)
        self.attrs_set = [list(map(int, a['values'])) for a in self.attrs]

    def _parse_attributes(self, attrs_input):
        """parses input attributes

        :warning: attrs_input is directly modified
        """
        for a in attrs_input:
            # TODO: handle name, type input?
            a['values'] = a['values'].split(',')
            if a['lambda'] == 'random':
                # generates [0,10] bounded scaled beta(2,2) random variable
                a['lambda'] = list(0 + 100 * self.prng.beta(2,2,size=len(a['values'])))
            if a['theta'] == 'random':
                # generates [-0.75,0.75] bounded scaled beta(2,2) random variable
                a['theta'] = list(-0.75 + 1.5 * self.prng.beta(2,2,size=len(a['values'])))
            if a['avg-revenue'] == 'random':
                # generates [30, 70] bounded scaled beta(2,2) random variable (after all attributes considered)
                # this makes avg rev per conversion 50.
                a['avg-revenue'] = list(30 / len(attrs_input) + 40 / len(attrs_input) * self.prng.beta(2,2,size=len(a['values'])))
            if a['prob-conversion'] == 'random':
                # generates [0, 0.15/len-attr] bounded scaled beta(2,2) r.v. to limit prob-conversion capped at 0.15
                a['prob-conversion'] = list(0 + 0.15 / len(attrs_input) * self.prng.beta(2,2,size=len(a['values'])))

        return attrs_input

    @staticmethod
    def get_revenue_sample(avg_revenue, prng=None, size=1):
        """
        log-normal revenue sample, with controlled average and prng
        :param avg_revenue: samples will have this as average
        :param prng: (optional) numpy.random.RandomState object
        :return: >0 sample
        """

        if avg_revenue <= 0:
            r = [0.0] * size
        else:
            if prng is None:
                prng = np.random.RandomState()
            # r = list(prng.lognormal(mean=0, sigma=np.sqrt(2 * np.log(avg_revenue)), size=size))
            r = list(prng.gamma(shape=4, scale=avg_revenue/4, size=size))    # lognormal too noisy, use gamma
        return [float(r_elem) for r_elem in r]

    @staticmethod
    def get_conversion(prob_conv, prng=None, size=1):
        if prng is None:
            prng = np.random.RandomState()
        r = list(prng.binomial(n=1, p=prob_conv, size=size))
        return [int(r_elem) for r_elem in r]

    def generate_sample(self):
        """generates auction arrivals, with all possible combination of attributes

        :return sample trajectory of auction arrivals

        The output can be fed into simulator so that auctions can take place
        """
        aucts = []
        for t in range(1, self.max_t+1):
            for attr in product(*self.attrs_set):
                auct = {}
                auct['attr'] = attr
                auct['iter'] = t
                auct['lambda'] = 0
                for ix, a_elem in enumerate(attr):
                    auct['lambda'] += self.attrs[ix]['lambda'][a_elem]
                auct['num_auct'] = self.prng.poisson(auct['lambda'])
                auct['theta'] = 0
                for ix, a_elem in enumerate(attr):
                    auct['theta'] += self.attrs[ix]['theta'][a_elem]
                auct['avg_revenue'] = 0
                for ix, a_elem in enumerate(attr):
                    auct['avg_revenue'] += self.attrs[ix]['avg-revenue'][a_elem]
                auct['prob_conversion'] = 0
                for ix, a_elem in enumerate(attr):
                    auct['prob_conversion'] += self.attrs[ix]['prob-conversion'][a_elem]
                aucts.append(auct)
            pass
        return aucts

if __name__ == "__main__":

    import pickle

    param, attrs = Auction.read_init_xlsx("auction_ini_01.xlsx")

    auction = Auction(param, attrs)
    gen = auction.generate_sample()

    pickle.dump(gen, open("auction_01.p", "wb"))
